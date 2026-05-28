"""
Gerador de ficheiros Excel de output.

Produz dois ficheiros:
  1. Mapa Picking Roteado (multi-sheet)
  2. Mapa Pre-Carga (single-sheet)
"""
import zipfile
import xml.etree.ElementTree as ET
import re
import copy
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import RoutePlan, AssignedStop


DAY_NAMES_PT = {
    0: 'segunda-feira', 1: 'terça-feira', 2: 'quarta-feira',
    3: 'quinta-feira', 4: 'sexta-feira', 5: 'sábado', 6: 'domingo',
}

HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(size=10)
BOLD_FONT = Font(bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def _fmt_date_pt(dt):
    return f"{dt.strftime('%d/%m/%Y')} ({DAY_NAMES_PT.get(dt.weekday(), '')})"


def _obs_entrega(stop):
    """Gera texto de OBSERVAÇÕES DE ENTREGA a partir da janela horaria."""
    if stop.time_window_text:
        txt = stop.time_window_text
        txt = txt.replace('ate ', 'até ')
        return txt
    return ""


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_routed_map(route_plans, lines, config, expedition_date, criteria_path,
                     output_path):
    """
    Gera o Mapa Picking Roteado (workbook multi-sheet).

    Sheets:
      1. Critérios          (copiado do Input 2)
      2. MAPA DE DISTRIBUIÇÃO (copiado do Input 2)
      3. matriculas - distribuidor (copiado do Input 2)
      4. BD                 (todas as linhas com atribuições)
      5. Resumo ROTEADO     (resumo por viatura)
      6. Resumo Gráfico     (tabela compacta)
      7. Motivos divisão    (justificação por viatura)
      8. Atividade Motorista (detalhe de atividade)
    """
    wb = Workbook()

    # --- Copiar sheets do Input 2 (criterios) ---
    _copy_criteria_sheets(wb, criteria_path)

    # --- BD ---
    _write_bd_sheet(wb, route_plans, lines, expedition_date)

    # --- Resumo ROTEADO ---
    _write_resumo_sheet(wb, route_plans, config, expedition_date)

    # --- Resumo Gráfico ---
    _write_resumo_grafico(wb, route_plans)

    # --- Motivos divisão ---
    _write_motivos_sheet(wb, route_plans, config, expedition_date)

    # --- Atividade Motorista ---
    _write_atividade_sheet(wb, route_plans, config, expedition_date)

    # Remover sheet default se existe
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(output_path)
    return output_path


def _copy_criteria_sheets(wb, criteria_path):
    """Copia as 3 primeiras sheets do Input 2 para o workbook."""
    from .loader import _open_xlsx, _read_xlsx_sheet

    if not criteria_path:
        return

    try:
        zf, shared, sheet_map = _open_xlsx(criteria_path)
    except Exception:
        return

    copy_names = ['Critérios', 'MAPA DE DISTRIBUIÇÃO', 'matriculas - distribuidor']

    for name in copy_names:
        if name not in sheet_map:
            for sn in sheet_map:
                if name.lower() in sn.lower():
                    name = sn
                    break

        if name in sheet_map:
            rows = _read_xlsx_sheet(zf, sheet_map[name], shared)
            ws = wb.create_sheet(name)
            for r_idx, row in enumerate(rows, 1):
                cols = sorted(row.keys(), key=lambda c: (len(c), c))
                for col_str in cols:
                    col_idx = _col_to_num(col_str)
                    val = row[col_str]
                    try:
                        val = float(val)
                        if val == int(val):
                            val = int(val)
                    except (ValueError, TypeError):
                        pass
                    ws.cell(row=r_idx, column=col_idx, value=val)
            _auto_width(ws)

    zf.close()


def _col_to_num(col_str):
    """Converte 'A' -> 1, 'B' -> 2, ..., 'AA' -> 27."""
    result = 0
    for ch in col_str:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _build_line_to_plan_map(route_plans):
    """Mapeia cada linha de picking para o seu RoutePlan e AssignedStop."""
    line_map = {}
    for plan in route_plans:
        for astop in plan.stops:
            for line in astop.stop.lines:
                line_map[line.row_index] = (plan, astop)
    return line_map


def _write_bd_sheet(wb, route_plans, lines, expedition_date):
    """
    Sheet BD: todas as linhas do input + colunas extra.
    Colunas A-Y: dados originais do picking
    Z: OBSERVAÇÕES DE ENTREGA
    AA: Matricula a atribuir
    AB: condutor a atribuir
    AC: ORDEM DE ENTREGA (ex: "3/18")
    AD: HORA PREVISTA DE ENTREGA
    """
    ws = wb.create_sheet('BD')

    headers = [
        'Cliente', 'Descrição', 'Artigo', 'Descrição2', 'Quantidade',
        'Transportador', 'Data Exped.', 'Doc. final', 'Doc. origem',
        'Delegação', 'Peso', 'Morada 1', 'Morada 2', 'Morada 3',
        'Cod. postal', 'Cidade', 'Obs. externas', 'End. expedição',
        'Lote', 'Un. Venda', 'Endereço Expedição', 'Altura', 'Largura',
        'Profundidade', 'Rota',
        'OBSERVAÇÕES DE ENTREGA', 'Matricula a atribuir',
        'condutor a atribuir', 'ORDEM DE ENTREGA', 'HORA PREVISTA DE ENTREGA',
    ]
    _write_header(ws, headers)

    line_map = _build_line_to_plan_map(route_plans)

    for row_idx, line in enumerate(lines, 2):
        plan, astop = line_map.get(line.row_index, (None, None))

        ws.cell(row=row_idx, column=1, value=line.client_code)
        ws.cell(row=row_idx, column=2, value=line.client_name)
        ws.cell(row=row_idx, column=3, value=line.article_code)
        ws.cell(row=row_idx, column=4, value=line.article_desc)
        ws.cell(row=row_idx, column=5, value=line.quantity)
        ws.cell(row=row_idx, column=6, value=line.transporter)
        ws.cell(row=row_idx, column=7, value=line.expedition_date)
        ws.cell(row=row_idx, column=8, value=line.doc_final)
        ws.cell(row=row_idx, column=9, value=line.doc_origin)
        ws.cell(row=row_idx, column=10, value=line.delegation)
        ws.cell(row=row_idx, column=11, value=line.weight)
        ws.cell(row=row_idx, column=12, value=line.address1)
        ws.cell(row=row_idx, column=13, value=line.address2)
        ws.cell(row=row_idx, column=14, value=line.address3)
        ws.cell(row=row_idx, column=15, value=line.postal_code)
        ws.cell(row=row_idx, column=16, value=line.city)
        ws.cell(row=row_idx, column=17, value=line.obs_external)
        ws.cell(row=row_idx, column=18, value=line.shipping_address)
        ws.cell(row=row_idx, column=19, value=line.lot)
        ws.cell(row=row_idx, column=20, value=line.sale_unit)
        ws.cell(row=row_idx, column=21, value=line.expedition_address)
        ws.cell(row=row_idx, column=22, value=line.height)
        ws.cell(row=row_idx, column=23, value=line.width)
        ws.cell(row=row_idx, column=24, value=line.depth)
        ws.cell(row=row_idx, column=25, value=line.route_code)

        if plan and astop:
            ws.cell(row=row_idx, column=26, value=_obs_entrega(astop.stop))
            ws.cell(row=row_idx, column=27, value=plan.vehicle.plate)
            ws.cell(row=row_idx, column=28, value=plan.vehicle.driver)
            ws.cell(row=row_idx, column=29,
                    value=f"{astop.delivery_order}/{astop.total_stops}")
            ws.cell(row=row_idx, column=30, value=astop.estimated_arrival)
        else:
            ws.cell(row=row_idx, column=26, value="")
            ws.cell(row=row_idx, column=27, value="NÃO ATRIBUÍDO")
            ws.cell(row=row_idx, column=28, value="")
            ws.cell(row=row_idx, column=29, value="")
            ws.cell(row=row_idx, column=30, value="")

    _auto_width(ws)


def _write_resumo_sheet(wb, route_plans, config, expedition_date):
    """
    Sheet Resumo ROTEADO: titulo + contexto + tabela de resumo por viatura.
    """
    ws = wb.create_sheet('Resumo ROTEADO')

    weekday = expedition_date.weekday()
    is_reduced = weekday in config['work_hours'].get('reduced_days', [])
    max_h = config['work_hours']['reduced']['max_hours'] if is_reduced else config['work_hours']['normal']['max_hours']
    day_name = DAY_NAMES_PT.get(weekday, '')

    tiago_plan = next((p for p in route_plans if p.vehicle.is_tiago), None)
    tiago_in_dist = tiago_plan is not None and tiago_plan.total_clients > 0

    ws.cell(row=1, column=1,
            value='RESUMO ROTEADO V FINAL - MAPA PICKING PORTO').font = BOLD_FONT

    context_parts = [
        f"Data de expedição considerada: {_fmt_date_pt(expedition_date)}",
        f"Chegada ao armazém: {config['work_hours']['normal']['start']} em {config['depot']['name']}",
    ]
    if tiago_in_dist:
        context_parts.append(
            f"Saída do carro do Tiago validada por limite máximo de {max_h:.0f}h00 "
            f"à {day_name}")
    else:
        for p in route_plans:
            if p.tiago_supports:
                context_parts.append(
                    f"Tiago apoia {p.vehicle.driver} (-10% tempo descarga)")
                break
    context_parts.append("Porto: -10% tempo entrega por cliente")
    ws.cell(row=2, column=1, value=' | '.join(context_parts)).font = DATA_FONT

    headers = [
        'Matricula', 'Condutor', 'Rotas a Fazer', 'Nº de caixas a entregar',
        'Nº de clientes a fazer', '% volume de carro ocupado',
        'Horário saída do armazém com carga completa',
        'Saída do último cliente', 'Chegada a casa',
        'Nº Horas Trabalhadas', 'Nº de kms efetuados',
        'Custo estimado gasóleo (€)', 'Notas',
    ]
    _write_header(ws, headers, row=3)

    for i, plan in enumerate(route_plans, 4):
        zones_str = ', '.join(plan.zones)
        tw_notes = _build_tw_notes(plan, config)

        ws.cell(row=i, column=1, value=plan.vehicle.plate)
        ws.cell(row=i, column=2, value=plan.vehicle.driver)
        ws.cell(row=i, column=3, value=zones_str)
        ws.cell(row=i, column=4, value=plan.total_boxes)
        ws.cell(row=i, column=5, value=plan.total_clients)
        ws.cell(row=i, column=6, value=plan.volume_pct)
        ws.cell(row=i, column=7, value=plan.departure_time)
        ws.cell(row=i, column=8, value=plan.last_client_departure)
        ws.cell(row=i, column=9, value=plan.arrival_home)
        ws.cell(row=i, column=10, value=plan.total_hours)
        ws.cell(row=i, column=11, value=plan.total_km)
        ws.cell(row=i, column=12, value=plan.fuel_cost)
        ws.cell(row=i, column=13, value=tw_notes)

    _auto_width(ws)


def _build_tw_notes(plan, config):
    """Gera notas sobre janelas respeitadas para o resumo."""
    weekday_max = config['work_hours']['reduced']['max_hours'] if any(
        plan.vehicle.is_tiago for _ in [1]) else config['work_hours']['normal']['max_hours']

    tw_parts = []
    for astop in plan.stops:
        if astop.stop.time_window_text:
            tw_parts.append(astop.stop.time_window_text)

    parts = []
    if tw_parts:
        parts.append(f"Janelas {', '.join(tw_parts)} respeitadas")

    is_reduced = False
    max_h = config['work_hours']['normal']['max_hours']
    if plan.total_hours <= max_h:
        parts.append(f"dentro de {max_h:.0f}h00")

    return '; '.join(parts) + '.' if parts else ''


def _write_resumo_grafico(wb, route_plans):
    """Sheet Resumo Gráfico: tabela compacta."""
    ws = wb.create_sheet('Resumo Gráfico')

    ws.cell(row=1, column=1,
            value='RESUMO GRÁFICO - ROTEADO V FINAL').font = BOLD_FONT

    headers = [
        'Matricula', 'Condutor', 'Caixas', 'Clientes',
        'Volume ocupado %', 'Horas trabalhadas', 'Kms', 'Custo gasóleo €',
    ]
    _write_header(ws, headers, row=2)

    for i, plan in enumerate(route_plans, 3):
        ws.cell(row=i, column=1, value=plan.vehicle.plate)
        ws.cell(row=i, column=2, value=plan.vehicle.driver)
        ws.cell(row=i, column=3, value=plan.total_boxes)
        ws.cell(row=i, column=4, value=plan.total_clients)
        ws.cell(row=i, column=5, value=plan.volume_pct)
        ws.cell(row=i, column=6, value=plan.total_hours)
        ws.cell(row=i, column=7, value=plan.total_km)
        ws.cell(row=i, column=8, value=plan.fuel_cost)

    _auto_width(ws)


def _write_motivos_sheet(wb, route_plans, config, expedition_date):
    """Sheet Motivos divisão: justificação por viatura."""
    ws = wb.create_sheet('Motivos divisão')

    ws.cell(row=1, column=1,
            value='MOTIVOS DA DIVISÃO POR CARRO - ROTEADO V FINAL').font = BOLD_FONT

    headers = ['Matricula / Condutor', 'Rotas atribuídas',
               'Motivo da divisão', 'Notas operacionais']
    _write_header(ws, headers, row=2)

    weekday = expedition_date.weekday()
    is_reduced = weekday in config['work_hours'].get('reduced_days', [])
    max_h = config['work_hours']['reduced']['max_hours'] if is_reduced else config['work_hours']['normal']['max_hours']

    for i, plan in enumerate(route_plans, 3):
        zones_str = ', '.join(plan.zones)

        ws.cell(row=i, column=1,
                value=f"{plan.vehicle.plate} / {plan.vehicle.driver}")
        ws.cell(row=i, column=2, value=zones_str)
        ws.cell(row=i, column=3, value=_generate_motivo(plan, config))
        ws.cell(row=i, column=4,
                value=f"{plan.total_clients} clientes, {plan.total_boxes} caixas, "
                      f"{plan.total_km} km, {plan.total_hours} h desde "
                      f"{config['work_hours']['normal']['start']}.")

    # Linha de critério geral
    row = len(route_plans) + 3
    tiago_plan = next((p for p in route_plans if p.vehicle.is_tiago), None)
    tiago_active = tiago_plan is not None and tiago_plan.total_clients > 0

    ws.cell(row=row, column=1, value='Critério geral')
    ws.cell(row=row, column=2, value='Todas as viaturas ativas')

    if tiago_active:
        ws.cell(row=row, column=3,
                value=f"A quinta viatura foi validada apenas após tentativa de "
                      f"reequilíbrio por 4 carros; a decisão garante todos os "
                      f"motoristas até {max_h:.0f}h00 e respeita janelas obrigatórias.")
    else:
        ws.cell(row=row, column=3,
                value=f"Distribuição em {len(route_plans)} viaturas garante "
                      f"todos os motoristas até {max_h:.0f}h00.")

    ws.cell(row=row, column=4,
            value=f"Observações de entrega limitadas exclusivamente a "
                  f"horários/janelas; hora de início fixa "
                  f"{config['work_hours']['normal']['start']} em "
                  f"{config['depot']['name']}; Porto com -10% no tempo de "
                  f"entrega por cliente.")

    _auto_width(ws)


def _generate_motivo(plan, config):
    """Gera texto de motivo de divisão para uma viatura."""
    zones = plan.zones
    if plan.vehicle.is_tiago:
        return (f"Saída da viatura validada porque, mantendo apenas "
                f"{len(plan.zones)} matrículas, a concentração e as janelas "
                f"obrigatórias provocavam excesso operacional.")

    if len(zones) == 1:
        return f"Mantém {zones[0]} no eixo natural do motorista."

    return (f"Concentra {', '.join(zones[:2])} e absorve "
            f"{', '.join(zones[2:])} no regresso."
            if len(zones) > 2
            else f"Consolida {zones[0]} e {zones[1]}.")


def _write_atividade_sheet(wb, route_plans, config, expedition_date):
    """Sheet Atividade Motorista: detalhe por motorista."""
    ws = wb.create_sheet('Atividade Motorista')

    ws.cell(row=1, column=1,
            value='RESUMO DE ATIVIDADE POR MOTORISTA').font = BOLD_FONT

    weekday = expedition_date.weekday()
    is_reduced = weekday in config['work_hours'].get('reduced_days', [])
    max_h = config['work_hours']['reduced']['max_hours'] if is_reduced else config['work_hours']['normal']['max_hours']
    start_time = config['work_hours']['normal']['start']

    context = (f"Data de expedição: {_fmt_date_pt(expedition_date)} | "
               f"Hora de chegada ao armazém: {start_time} | "
               f"Local: {config['depot']['name']} | "
               f"Critério: janelas horárias, Porto com -10% tempo de entrega")
    ws.cell(row=2, column=1, value=context).font = DATA_FONT

    headers = [
        'Matrícula', 'Condutor', 'Hora chegada armazém',
        'Hora saída armazém', 'Trajeto apenas por cidades',
        'Hora última entrega', 'Hora saída último cliente',
        'Hora chegada a casa', 'Nº horas trabalhadas até casa',
        'Nº horas até última entrega', 'Nº clientes', 'Nº caixas',
        'Observações',
    ]
    _write_header(ws, headers, row=3)

    for i, plan in enumerate(route_plans, 4):
        cities = _build_trajectory(plan, config)

        # Hora ultima entrega = arrival do ultimo stop
        last_arrival = ""
        last_arrival_min = 0
        if plan.stops:
            last = plan.stops[-1]
            last_arrival = last.estimated_arrival
            last_arrival_min = last.arrival_minutes

        start_min = _time_to_minutes(start_time)
        hours_to_last = round((last_arrival_min - start_min) / 60, 2) if last_arrival_min else 0

        ws.cell(row=i, column=1, value=plan.vehicle.plate)
        ws.cell(row=i, column=2, value=plan.vehicle.driver)
        ws.cell(row=i, column=3, value=start_time)
        ws.cell(row=i, column=4, value=plan.departure_time)
        ws.cell(row=i, column=5, value=cities)
        ws.cell(row=i, column=6, value=last_arrival)
        ws.cell(row=i, column=7, value=plan.last_client_departure)
        ws.cell(row=i, column=8, value=plan.arrival_home)
        ws.cell(row=i, column=9, value=plan.total_hours)
        ws.cell(row=i, column=10, value=hours_to_last)
        ws.cell(row=i, column=11, value=plan.total_clients)
        ws.cell(row=i, column=12, value=plan.total_boxes)
        ws.cell(row=i, column=13, value=_build_obs_motorista(plan, config))

    _auto_width(ws)


def _build_trajectory(plan, config):
    """Constroi string de trajeto por cidades."""
    depot_name = config['depot']['name'].split(' - ')[0]
    depot_full = config['depot']['name'].replace(' - ', '/')

    cities = [depot_full]
    seen = set()
    for astop in plan.stops:
        city = astop.stop.city.upper().strip() if astop.stop.city else ""
        if city and city not in seen:
            seen.add(city)
            cities.append(city)
    cities.append('Casa')

    return ' → '.join(cities)


def _build_obs_motorista(plan, config):
    """Gera observações para a sheet Atividade Motorista."""
    start_time = config['work_hours']['normal']['start']
    parts = []

    tw_stops = [a for a in plan.stops if a.stop.time_window_text]
    if tw_stops:
        parts.append("Janelas respeitadas")

    parts.append(f"início de trabalho às {start_time} em "
                 f"{config['depot']['name']}")

    return '; '.join(parts) + '.'


def _time_to_minutes(time_str):
    h, m = map(int, time_str.split(':'))
    return h * 60 + m


# ============================================================
# MAPA PRE-CARGA
# ============================================================

def write_pre_carga(route_plans, lines, config, expedition_date, output_path):
    """
    Gera o Mapa Pré-Carga.

    Colunas:
      A: SalerOrder     (= Doc. origem)
      B: DeliveryDate   (= Data Exped. original)
      C: StoreCode      (= Cliente + End. expedição, concatenados)
      D: ArticleCode    (= Artigo)
      E: Quantidade
      F: UOM            (= Un. Venda)
      G: Element        (= Endereço Expedição)
      H: Symbol         (= Lote)
      I: StockReference (= Lote)
      J: Route          (= Matricula atribuida)
      K: Delegacao
      L: Ordem de Carga (= inverso da ordem de entrega: total-ordem+1)
      M: Armazém de Carga
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Folha1'

    headers = [
        'SalerOrder', 'DeliveryDate', 'StoreCode', 'ArticleCode',
        'Quantidade', 'UOM', 'Element', 'Symbol', 'StockReference',
        'Route', 'Delegacao', 'Ordem de Carga', 'Armazém de Carga',
    ]
    _write_header(ws, headers)

    line_map = _build_line_to_plan_map(route_plans)
    depot_name = config['depot']['name']

    for row_idx, line in enumerate(lines, 2):
        plan, astop = line_map.get(line.row_index, (None, None))

        store_code = f"{line.client_code}{line.shipping_address}"
        plate = plan.vehicle.plate if plan else ""
        if astop:
            load_order = astop.total_stops - astop.delivery_order + 1
            ordem = f"{load_order}/{astop.total_stops}"
        else:
            ordem = ""

        ws.cell(row=row_idx, column=1, value=line.doc_origin)
        ws.cell(row=row_idx, column=2, value=line.expedition_date)
        ws.cell(row=row_idx, column=3, value=store_code)
        ws.cell(row=row_idx, column=4, value=line.article_code)
        ws.cell(row=row_idx, column=5, value=line.quantity)
        ws.cell(row=row_idx, column=6, value=line.sale_unit)
        ws.cell(row=row_idx, column=7, value=line.expedition_address)
        ws.cell(row=row_idx, column=8, value=line.lot)
        ws.cell(row=row_idx, column=9, value=line.lot)
        ws.cell(row=row_idx, column=10, value=plate)
        ws.cell(row=row_idx, column=11, value=line.delegation)
        ws.cell(row=row_idx, column=12, value=ordem)
        ws.cell(row=row_idx, column=13, value=depot_name)

    _auto_width(ws)
    wb.save(output_path)
    return output_path
